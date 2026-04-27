"""ADR-014 Task 3: plain-text extraction from xlsx and docx files.

Produces flat text suitable for narrative-LLM context.  Output format
follows ADR-014 §6 exactly:

xlsx → ``[Excel-файл: name]`` + per-sheet TSV-style rows.
docx → ``[Word-файл: name]`` + paragraphs and table rows in body order.
"""

from __future__ import annotations

import logging
from pathlib import Path

logger = logging.getLogger("analysis.media_extract.office")


class OfficeParseError(Exception):
    """Ошибка парсинга Office-документа."""


def extract_xlsx(path: Path) -> str:
    """Parse an Excel file (xlsx) into flat text.

    Raises:
        FileNotFoundError: if the file does not exist.
        OfficeParseError: on read failure (corrupt / unsupported format).
    """
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise OfficeParseError(f"Failed to read xlsx '{path.name}': {exc}") from exc

    logger.info("Parsing xlsx: %s, sheets: %d", path.name, len(wb.sheetnames))

    lines: list[str] = [f"[Excel-файл: {path.name}]"]

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_texts: list[str] = []
            for row in ws.iter_rows():
                cells = [str(cell.value) if cell.value is not None else "" for cell in row]
                if all(c == "" for c in cells):
                    continue
                row_texts.append("\t".join(cells))

            if not row_texts:
                lines.append(f'Лист "{sheet_name}": (пусто)')
            else:
                lines.append(f'Лист "{sheet_name}":')
                lines.extend(row_texts)
    finally:
        wb.close()

    return "\n".join(lines)


def extract_docx(path: Path) -> str:
    """Parse a Word file (docx) into flat text.

    Iterates the document body in element order so paragraphs and tables
    appear in their original sequence.

    Raises:
        FileNotFoundError: if the file does not exist.
        OfficeParseError: on read failure.
    """
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    try:
        import docx
        from docx.oxml.ns import qn
        from docx.table import Table
        from docx.text.paragraph import Paragraph

        document = docx.Document(str(path))
    except Exception as exc:
        raise OfficeParseError(f"Failed to read docx '{path.name}': {exc}") from exc

    logger.info("Parsing docx: %s", path.name)

    lines: list[str] = [f"[Word-файл: {path.name}]"]

    for block in document.element.body:
        tag = block.tag
        if tag == qn("w:p"):
            text = Paragraph(block, document).text.strip()
            if text:
                lines.append(text)
        elif tag == qn("w:tbl"):
            table = Table(block, document)
            for row in table.rows:
                cell_texts = [cell.text.strip() for cell in row.cells]
                lines.append("\t".join(cell_texts))

    return "\n".join(lines)
