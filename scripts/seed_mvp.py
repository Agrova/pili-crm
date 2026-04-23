"""MVP seed — load CRM_late.xlsx + tg_scan_results.json into Postgres.

Idempotent: wipes derived tables (suppliers / products / orders / items /
customers) and re-inserts from the Excel source. Enables pg_trgm once.

Usage:
    python3 scripts/seed_mvp.py
"""

from __future__ import annotations

import asyncio
import json
import sys
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.catalog.models import CatalogProduct, CatalogProductListing, CatalogSupplier
from app.database import async_session_factory
from app.orders.models import (
    OrdersCustomer,
    OrdersCustomerProfile,
    OrdersOrder,
    OrdersOrderItem,
    OrdersOrderItemStatus,
    OrdersOrderStatus,
)
from app.orders.service import derive_order_status
from app.pricing import models as _pricing_models  # noqa: F401  register mappers
from app.warehouse.models import WarehouseStockItem

SEED_DIR = Path(__file__).resolve().parent.parent / "data" / "seed"
EXCEL_PATH = SEED_DIR / "CRM_late.xlsx"
TG_SCAN_PATH = SEED_DIR / "tg_scan_results.json"

# All orders start as draft; status is computed from items after they're loaded.
ORDER_STATUS_MAP = {
    "Новый": OrdersOrderStatus.draft,
    "Заказан у поставщика": OrdersOrderStatus.draft,
    "В пути": OrdersOrderStatus.draft,
    "Получен": OrdersOrderStatus.draft,
    "Передан клиенту": OrdersOrderStatus.draft,
    "Закрыт": OrdersOrderStatus.draft,
}

ITEM_STATUS_MAP = {
    "Нужно заказать": OrdersOrderItemStatus.pending,
    "Заказан": OrdersOrderItemStatus.ordered,
    "Заказан у поставщика": OrdersOrderItemStatus.ordered,
    "Получен": OrdersOrderItemStatus.delivered,
    "Передан клиенту": OrdersOrderItemStatus.delivered,
}

UNKNOWN_SUPPLIER = "Unknown"


@dataclass
class SeedReport:
    suppliers: int = 0
    products: int = 0
    customers: int = 0
    orders: int = 0
    items: int = 0
    telegram_enriched: int = 0


_NULL_PLACEHOLDERS = {"", "-", "—", "–", "n/a", "N/A"}


def _norm(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in _NULL_PLACEHOLDERS:
        return None
    return s or None


def _dec(v: object) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


async def _wipe(s: AsyncSession) -> None:
    # FK order: stock → listings → products → suppliers; items → orders → customers
    await s.execute(delete(WarehouseStockItem))
    await s.execute(delete(OrdersOrderItem))
    await s.execute(delete(OrdersOrder))
    await s.execute(delete(OrdersCustomerProfile))
    await s.execute(delete(OrdersCustomer))
    await s.execute(delete(CatalogProductListing))
    await s.execute(delete(CatalogProduct))
    await s.execute(delete(CatalogSupplier))


async def _ensure_pg_trgm(s: AsyncSession) -> None:
    await s.execute(text("CREATE EXTENSION IF NOT EXISTS pg_trgm"))


async def _load_suppliers(s: AsyncSession, wb) -> dict[str, int]:
    ws = wb["Позиции заказа"]
    names: set[str] = {UNKNOWN_SUPPLIER}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = _norm(row[13]) or UNKNOWN_SUPPLIER
        names.add(name)

    result: dict[str, int] = {}
    for name in sorted(names):
        supplier = CatalogSupplier(name=name)
        s.add(supplier)
        await s.flush()
        result[name] = supplier.id
    return result


async def _load_products(
    s: AsyncSession, wb, supplier_ids: dict[str, int]
) -> dict[tuple[str, str], int]:
    ws = wb["Позиции заказа"]
    seen: dict[tuple[str, str], tuple[str | None, Decimal | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = _norm(row[2])
        if not name:
            continue
        supplier = _norm(row[13]) or UNKNOWN_SUPPLIER
        sku = _norm(row[3])
        weight = _dec(row[6])
        key = (supplier, name)
        if key not in seen:
            seen[key] = (sku, weight)

    result: dict[tuple[str, str], int] = {}
    for (supplier, name), (sku, weight) in seen.items():
        product = CatalogProduct(
            name=name,
            sku=sku,
            declared_weight=weight if weight and weight > 0 else None,
        )
        s.add(product)
        await s.flush()
        listing = CatalogProductListing(
            product_id=product.id,
            source_id=supplier_ids[supplier],
            sku_at_source=sku,
            is_primary=True,
        )
        s.add(listing)
        await s.flush()
        result[(supplier, name)] = product.id
    return result


async def _load_customers(s: AsyncSession, wb) -> dict[str, int]:
    ws = wb["Клиенты"]
    result: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        excel_id = _norm(row[0])
        name = _norm(row[1])
        if not excel_id or not name:
            continue
        telegram = _norm(row[2])
        phone = _norm(row[3])
        email = _norm(row[6])
        # Contact-check constraint: email OR phone OR telegram_id required.
        # Use placeholder telegram_id tied to Excel ID when all are empty.
        if not (telegram or phone or email):
            telegram = f"@excel_{excel_id}"
        customer = OrdersCustomer(
            name=name,
            email=email,
            phone=phone,
            telegram_id=telegram,
        )
        s.add(customer)
        await s.flush()
        result[excel_id] = customer.id
    return result


async def _load_orders(
    s: AsyncSession, wb, customer_ids: dict[str, int]
) -> dict[str, int]:
    ws = wb["Заказы"]
    result: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_code = _norm(row[0])
        excel_customer_id = _norm(row[1])
        status_raw = _norm(row[4])
        total = _dec(row[6])
        if not order_code or not excel_customer_id:
            continue
        if excel_customer_id not in customer_ids:
            print(f"  ! order {order_code}: unknown customer {excel_customer_id}")
            continue
        status = ORDER_STATUS_MAP.get(status_raw or "", OrdersOrderStatus.draft)
        order = OrdersOrder(
            customer_id=customer_ids[excel_customer_id],
            status=status,
            total_price=total,
            currency="RUB",
        )
        s.add(order)
        await s.flush()
        result[order_code] = order.id
    return result


async def _load_items(
    s: AsyncSession,
    wb,
    order_ids: dict[str, int],
    product_ids: dict[tuple[str, str], int],
) -> int:
    ws = wb["Позиции заказа"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_code = _norm(row[0])
        product_name = _norm(row[2])
        if not order_code or not product_name:
            continue
        if order_code not in order_ids:
            continue
        supplier = _norm(row[13]) or UNKNOWN_SUPPLIER
        key = (supplier, product_name)
        if key not in product_ids:
            continue
        qty = _dec(row[7]) or Decimal("1")
        unit_price = _dec(row[10])
        status_raw = _norm(row[12])
        status = ITEM_STATUS_MAP.get(
            status_raw or "", OrdersOrderItemStatus.pending
        )
        item = OrdersOrderItem(
            order_id=order_ids[order_code],
            product_id=product_ids[key],
            quantity=qty,
            unit_price=unit_price,
            status=status,
        )
        s.add(item)
        count += 1
    await s.flush()
    return count


async def _derive_all_order_statuses(
    s: AsyncSession, order_ids: dict[str, int]
) -> None:
    """Compute and persist order statuses from item statuses via derivation rule."""
    # Bulk-load all item statuses grouped by order_id
    rows = (
        await s.execute(
            select(OrdersOrderItem.order_id, OrdersOrderItem.status)
        )
    ).all()

    statuses_by_order: dict[int, list[str]] = {}
    for order_id, status in rows:
        statuses_by_order.setdefault(order_id, []).append(str(status))

    orders = (
        await s.execute(
            select(OrdersOrder).where(
                OrdersOrder.id.in_(order_ids.values())
            )
        )
    ).scalars().all()

    for order in orders:
        item_statuses = statuses_by_order.get(order.id, [])
        if not item_statuses:
            continue
        new_status = derive_order_status(item_statuses)
        order.status = new_status  # type: ignore[assignment]

    await s.flush()


async def _enrich_telegram(s: AsyncSession) -> int:
    if not TG_SCAN_PATH.exists():
        return 0
    data = json.loads(TG_SCAN_PATH.read_text(encoding="utf-8"))
    clients = [e for e in data if e.get("category") == "client"]
    customers = (await s.execute(select(OrdersCustomer))).scalars().all()
    enriched = 0
    for cust in customers:
        if cust.telegram_id and not cust.telegram_id.startswith("@excel_"):
            continue
        last_name = cust.name.split()[-1].lower()
        full = cust.name.lower()
        for entry in clients:
            cname = (entry.get("client_name") or "").lower()
            tname = (entry.get("name") or "").lower()
            if not cname:
                continue
            if last_name in cname or full in cname or cname in full:
                handle = entry.get("name") or entry.get("client_name")
                if not handle:
                    continue
                new_tg = f"@{handle.strip().replace(' ', '_')}"
                exists = (
                    await s.execute(
                        select(OrdersCustomer).where(
                            OrdersCustomer.telegram_id == new_tg
                        )
                    )
                ).scalar_one_or_none()
                if exists and exists.id != cust.id:
                    continue
                cust.telegram_id = new_tg
                enriched += 1
                break
            _ = tname
    return enriched


async def main() -> SeedReport:
    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(EXCEL_PATH, data_only=True)
    report = SeedReport()

    async with async_session_factory() as s, s.begin():
        await _ensure_pg_trgm(s)
        await _wipe(s)

        supplier_ids = await _load_suppliers(s, wb)
        report.suppliers = len(supplier_ids)

        product_ids = await _load_products(s, wb, supplier_ids)
        report.products = len(product_ids)

        customer_ids = await _load_customers(s, wb)
        report.customers = len(customer_ids)

        order_ids = await _load_orders(s, wb, customer_ids)
        report.orders = len(order_ids)

        report.items = await _load_items(s, wb, order_ids, product_ids)

        await _derive_all_order_statuses(s, order_ids)

        report.telegram_enriched = await _enrich_telegram(s)

    print("=" * 40)
    print("SEED REPORT")
    print("=" * 40)
    print(f"  suppliers:         {report.suppliers}")
    print(f"  products:          {report.products}")
    print(f"  customers:         {report.customers}")
    print(f"  orders:            {report.orders}")
    print(f"  order items:       {report.items}")
    print(f"  telegram enriched: {report.telegram_enriched}")
    return report


if __name__ == "__main__":
    asyncio.run(main())
