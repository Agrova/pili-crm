"""Parsing-level tests for the MVP seed script (no DB)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.seed_mvp import ITEM_STATUS_MAP, ORDER_STATUS_MAP

EXCEL = Path(__file__).resolve().parent.parent / "data" / "seed" / "CRM_late.xlsx"


@pytest.fixture(scope="module")
def wb():
    if not EXCEL.exists():
        pytest.skip(f"{EXCEL} not staged — run scripts/seed_mvp.py locally first")
    return load_workbook(EXCEL, data_only=True)


def _rows(ws, start=2):
    return [r for r in ws.iter_rows(min_row=start, values_only=True) if r[0]]


def test_customer_count(wb):
    assert len(_rows(wb["Клиенты"])) == 36


def test_order_count(wb):
    assert len(_rows(wb["Заказы"])) == 62


def test_item_count(wb):
    assert len(_rows(wb["Позиции заказа"])) == 133


def test_order_status_map_covers_sheet(wb):
    seen = {r[4] for r in _rows(wb["Заказы"]) if r[4]}
    missing = seen - ORDER_STATUS_MAP.keys()
    assert not missing, f"unmapped order statuses: {missing}"


def test_item_status_map_covers_sheet(wb):
    seen = {r[12] for r in _rows(wb["Позиции заказа"]) if r[12]}
    missing = seen - ITEM_STATUS_MAP.keys()
    assert not missing, f"unmapped item statuses: {missing}"


def test_pending_item_tally(wb):
    pending = {"Нужно заказать", "Заказан", "Заказан у поставщика"}
    n = sum(1 for r in _rows(wb["Позиции заказа"]) if r[12] in pending)
    assert n == 108
